"""
purity_hashtag.py — Chạy trong folder purity_get_hashtag/

Tính purity để chọn hashtag đặc trưng theo ngành và theo brand.

─── Công thức ──────────────────────────────────────────────────────────────────
  n(h, g)        = count_unique_video của hashtag h trong group g
  N(h)           = Σ n(h, g')  [tổng qua TẤT CẢ group cùng scope]

  purity(h, g)   = n(h, g) / N(h)
                   1.0  → h chỉ xuất hiện trong 1 group → đặc trưng tuyệt đối
                   ~0   → h dàn đều mọi group → generic, bỏ qua

  disc_score     = n(h, g) × purity(h, g)²
                   → ưu tiên hashtag vừa phổ biến, vừa đặc trưng

─── Scope tính N(h) ─────────────────────────────────────────────────────────────
  Category dict  : N(h) = tổng qua TẤT CẢ ngành  (hashtag beauty vs thời trang)
  Brand dict     : N(h) = tổng trong CÙNG NGÀNH   (tránh nhiễu chéo ngành)

─── Input ───────────────────────────────────────────────────────────────────────
  Category count hashtag.xlsx   — (category, hashtag, count_unique_video)
  Top Hashtag Per Brand In Cate.xlsx — (h.category, h.brand, hashtag, count_unique_video)

─── Output ──────────────────────────────────────────────────────────────────────
  output_category_hashtags.xlsx — cate | brand | hashtag_count | list_hashtag
  output_brand_hashtags.xlsx    — cate | brand | hashtag_count | list_hashtag
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ─── Đường dẫn ────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
CATE_FILE = BASE / "Category count hashtag.xlsx"
BRAND_FILE = BASE / "Top Hashtag Per Brand In Cate.xlsx"
OUT_CATE = BASE / "output_category_hashtags.xlsx"
OUT_BRAND = BASE / "output_brand_hashtags.xlsx"

# ─── Tham số — chỉnh tại đây ─────────────────────────────────────────────────
MIN_PURITY_CATEGORY = 0.5    # purity tối thiểu để giữ hashtag (category)
MIN_PURITY_BRAND    = 0.6    # purity tối thiểu (brand) — ngưỡng cao hơn
MIN_COUNT_CATEGORY  = 30     # số video tối thiểu (category)
MIN_COUNT_BRAND     = 30     # số video tối thiểu (brand)
TOP_N_CATEGORY      = 200    # số hashtag giữ lại / ngành
TOP_N_BRAND         = 30     # số hashtag giữ lại / brand
SEP                 = ", "   # ký tự phân cách trong list_hashtag

NOISE_BRANDS = {"Unknown", "Không Có"}

# Ngành target cho category dict (None = lấy tất cả)
TARGET_CATEGORIES: list[str] | None = [
    "Chăm sóc sắc đẹp & Chăm sóc cá nhân",
    "Trang phục nữ & Đồ lót",
    "Trang phục nam & Đồ lót",
    "Đồ ăn & Đồ uống",
    "Trẻ sơ sinh & thai sản",
    "Đồ gia dụng",
    "Phụ kiện thời trang",
]

VERSION = "v2026.07"


# ─── Purity core ──────────────────────────────────────────────────────────────

def add_purity(df: pd.DataFrame) -> pd.DataFrame:
    """
    Thêm cột purity và disc_score vào df.
    N(h) = tổng count_unique_video của hashtag h trong df (toàn bộ scope đã truyền vào).
    Gọi hàm này sau khi đã scope đúng (ví dụ: lọc 1 ngành).
    """
    df = df.copy()
    total = df.groupby("hashtag")["count_unique_video"].transform("sum")
    df["purity"] = df["count_unique_video"] / total
    df["disc_score"] = df["count_unique_video"] * df["purity"] ** 2
    return df


# ─── Category dict ────────────────────────────────────────────────────────────

def build_category_result(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Input : (category, hashtag, count_unique_video) — toàn bộ ngành.
    Output: (cate, brand, hashtag_count, list_hashtag) — 1 dòng / ngành.

    Purity tính global: N(h) = tổng qua TẤT CẢ ngành.
    Chọn top-N theo disc_score, lọc min_purity & min_count.
    """
    df = df_raw.copy()
    if TARGET_CATEGORIES:
        df = df[df["category"].isin(TARGET_CATEGORIES)]

    scored = add_purity(df)
    filtered = scored[
        (scored["purity"] >= MIN_PURITY_CATEGORY) &
        (scored["count_unique_video"] >= MIN_COUNT_CATEGORY)
    ]

    rows = []
    for cate, grp in filtered.groupby("category"):
        top = grp.nlargest(TOP_N_CATEGORY, "disc_score")["hashtag"].tolist()
        rows.append({
            "cate": cate,
            "brand": "",
            "hashtag_count": len(top),
            "list_hashtag": SEP.join(top),
        })
    return pd.DataFrame(rows).sort_values("cate").reset_index(drop=True)


# ─── Brand dict ───────────────────────────────────────────────────────────────

def build_brand_result(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Input : (h.category, h.brand, hashtag, count_unique_video).
    Output: (cate, brand, hashtag_count, list_hashtag) — 1 dòng / (ngành, brand).

    Purity tính within-category: N(h) = tổng qua các brand CÙNG NGÀNH.
    """
    df = df_raw[~df_raw["h.brand"].isin(NOISE_BRANDS)].copy()
    df = df.rename(columns={"h.category": "category", "h.brand": "brand"})

    parts = []
    for _, grp in df.groupby("category"):
        parts.append(add_purity(grp))
    scored = pd.concat(parts, ignore_index=True)

    filtered = scored[
        (scored["purity"] >= MIN_PURITY_BRAND) &
        (scored["count_unique_video"] >= MIN_COUNT_BRAND)
    ]

    rows = []
    for (cate, brand), grp in filtered.groupby(["category", "brand"]):
        top = grp.nlargest(TOP_N_BRAND, "disc_score")["hashtag"].tolist()
        if top:
            rows.append({
                "cate": cate,
                "brand": brand,
                "hashtag_count": len(top),
                "list_hashtag": SEP.join(top),
            })
    return pd.DataFrame(rows).sort_values(["cate", "brand"]).reset_index(drop=True)


# ─── Excel writer ─────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", start_color="1F3864")   # xanh đậm
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ROW_FONT     = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="EEF2F7")   # xám nhạt xen kẽ
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN    = Alignment(vertical="top")


def _write_result_sheet(ws, df: pd.DataFrame, col_widths: dict[str, int]) -> None:
    """Ghi sheet result với header + data + màu xen kẽ."""
    headers = df.columns.tolist()
    ws.append(headers)

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = TOP_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    for row_idx, row in df.iterrows():
        ws.append(row.tolist())
        excel_row = row_idx + 2
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = ROW_FONT
            if fill:
                cell.fill = fill
            cell.alignment = WRAP_ALIGN if col_name == "list_hashtag" else TOP_ALIGN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_readme_sheet(ws, title: str, scope_note: str) -> None:
    """Sheet readme giải thích thuật toán và các cột."""
    readme_font  = Font(name="Arial", size=10)
    title_font   = Font(name="Arial", bold=True, size=13, color="1F3864")
    section_font = Font(name="Arial", bold=True, size=11, color="2E5496")
    header_fill  = PatternFill("solid", start_color="D9E1F2")

    rows = [
        (title, "title"),
        ("", "blank"),
        ("THUAT TOAN PURITY", "section"),
        ("n(h, g)", "n(h, g) = so video cua group g co hashtag h"),
        ("N(h)", f"N(h) = tong n(h, g') qua tat ca group  [{scope_note}]"),
        ("purity(h, g)", "n(h, g) / N(h)   |  1.0 = chi xuat hien trong 1 group (dac trung tuyet doi)"),
        ("disc_score", "n(h,g) x purity^2  |  penalty nang hashtag generic; chon top-N theo chi so nay"),
        ("", "blank"),
        ("THAM SO", "section"),
        ("MIN_PURITY", str(MIN_PURITY_CATEGORY if "ategory" in title else MIN_PURITY_BRAND)),
        ("MIN_COUNT",  str(MIN_COUNT_CATEGORY  if "ategory" in title else MIN_COUNT_BRAND)),
        ("TOP_N",      str(TOP_N_CATEGORY       if "ategory" in title else TOP_N_BRAND)),
        ("Version",    VERSION),
        ("", "blank"),
        ("CAC COT OUTPUT", "section"),
        ("cate",          "Ten nganh hang"),
        ("brand",         "Ten brand  (trong file category: de trong)"),
        ("hashtag_count", "So hashtag da chon cho cap (cate, brand) nay"),
        ("list_hashtag",  f"Danh sach hashtag phan cach '{SEP}', sap xep theo disc_score giam dan"),
        ("", "blank"),
        ("LUU Y", "section"),
        ("Truncation",    "File brand chi co top 50 hashtag/brand -> N(h) co the thap hon thuc te"),
        ("Scope brand",   "Purity tinh within-category: tranh bi nhieu cheo nganh"),
        ("Noise brands",  f"Da loai: {', '.join(sorted(NOISE_BRANDS))}"),
    ]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    for label, value in rows:
        if value == "title":
            ws.append([label, ""])
            ws["A1"].font = title_font
            ws["A1"].fill = header_fill
        elif value == "blank":
            ws.append(["", ""])
        elif value == "section":
            ws.append([label, ""])
            r = ws.max_row
            ws.cell(row=r, column=1).font = section_font
        else:
            ws.append([label, value])
            r = ws.max_row
            ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)
            ws.cell(row=r, column=2).font = readme_font
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    ws.sheet_view.showGridLines = False


def save_excel(df: pd.DataFrame, out_path: Path, title: str, scope_note: str) -> None:
    col_widths = {"cate": 45, "brand": 35, "hashtag_count": 15, "list_hashtag": 100}
    wb = Workbook()
    ws_result = wb.active
    ws_result.title = "result"
    _write_result_sheet(ws_result, df, col_widths)

    ws_readme = wb.create_sheet("readme")
    _write_readme_sheet(ws_readme, title, scope_note)

    wb.save(out_path)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Reading: {CATE_FILE.name}")
    df_cate_raw = pd.read_excel(CATE_FILE, sheet_name="CustomQuery")
    print(f"  {len(df_cate_raw):,} rows | {df_cate_raw['category'].nunique()} categories")

    print(f"Reading: {BRAND_FILE.name}")
    df_brand_raw = pd.read_excel(BRAND_FILE, sheet_name="CustomQuery")
    df_brand_raw = df_brand_raw[~df_brand_raw["h.brand"].isin(NOISE_BRANDS)]
    print(f"  {len(df_brand_raw):,} rows | {df_brand_raw['h.category'].nunique()} categories | {df_brand_raw['h.brand'].nunique()} brands")

    # ── Category ─────────────────────────────────────────────────────────────
    df_cate_out = build_category_result(df_cate_raw)
    save_excel(
        df_cate_out, OUT_CATE,
        title="Category Hashtag Dict — Purity Algorithm",
        scope_note="N(h) tinh qua TAT CA nganh (category purity)",
    )
    print(f"\n[category] -> {OUT_CATE.name}")
    print(f"  {'Nganh':<50} {'hashtags':>8}")
    print(f"  {'-'*58}")
    for _, r in df_cate_out.iterrows():
        print(f"  {str(r['cate'])[:50]:<50} {r['hashtag_count']:>8}")

    # ── Brand ─────────────────────────────────────────────────────────────────
    df_brand_out = build_brand_result(df_brand_raw)
    save_excel(
        df_brand_out, OUT_BRAND,
        title="Brand Hashtag Dict — Purity Algorithm (within-category)",
        scope_note="N(h) tinh trong CUNG NGANH (brand purity, tranh nhieu cheo nganh)",
    )
    print(f"\n[brand]    -> {OUT_BRAND.name}")
    print(f"  {'Nganh':<45} {'Brand':<35} {'tags':>5}")
    print(f"  {'-'*85}")
    for _, r in df_brand_out.iterrows():
        print(f"  {str(r['cate'])[:45]:<45} {str(r['brand'])[:35]:<35} {r['hashtag_count']:>5}")

    print(f"\nDone. {len(df_cate_out)} nganh | {len(df_brand_out)} (nganh x brand) pairs")


if __name__ == "__main__":
    main()
